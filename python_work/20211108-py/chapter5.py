from random import *
from openpyxl import Workbook


def dxlsx(filename):
    wb = Workbook()
    ws = wb.active

    ws.append(["국어", "영어", "수학"])
    for i in range(1, 11):
        ws.append([randint(0, 100), randint(0, 100), randint(0, 100)])

    colran = ws["B:C"]

    for col in colran:
        for cell in col:
            print(cell.value)
    print()

    for column in ws.iter_cols():
        print(column[0].value)

    for row in ws.iter_rows():
        print(row[2].value)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            print(ws.cell(column=col, row=row).value)

    wb.save(filename + ".xlsx")
    wb.close()
