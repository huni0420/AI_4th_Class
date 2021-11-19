from openpyxl import load_workbook, Workbook


class StudentExcel:

    def __init__(self):
        pass

    def appendrow(self, num, adc, que1, que2, mid, las, pro, tot):
        wb = load_workbook('student_scores.xlsx')
        ws = wb.active
        a = int(adc)
        if a > 4:
            if tot > 90:
                self.t ="A"
            elif tot > 80:
                self.t ="B"
            elif tot > 70:
                self.t="C"
            elif tot <= 70:
                self.t="D"
        elif a <= 4:
            self.t = "F"

        ws.cell(row=int(num)+1, column=1).value = str(num)
        ws.cell(row=int(num)+1, column=2).value = str(adc)
        ws.cell(row=int(num)+1, column=3).value = str(que1)
        ws.cell(row=int(num)+1, column=4).value = str(que2)
        ws.cell(row=int(num)+1, column=5).value = str(mid)
        ws.cell(row=int(num)+1, column=6).value = str(las)
        ws.cell(row=int(num)+1, column=7).value = str(pro)
        ws.cell(row=int(num)+1, column=8).value = str(tot)
        ws.cell(row=int(num)+1, column=9).value = self.t

        wb.save('student_scores.xlsx')
        wb.close()

    def correction(self, num, adc, que1, que2, mid, las, pro, tot, c):
        wb = load_workbook('student_scores.xlsx')
        ws = wb.active


        ws.cell(row=int(num), column=1).value = num
        ws.cell(row=int(num), column=2).value = adc
        ws.cell(row=int(num), column=3).value = que1
        ws.cell(row=int(num), column=4).value = que2
        ws.cell(row=int(num), column=5).value = mid
        ws.cell(row=int(num), column=6).value = las
        ws.cell(row=int(num), column=7).value = pro
        ws.cell(row=int(num), column=8).value = tot
        ws.cell(row=int(num), column=9).value = c

        wb.save('student_scores.xlsx')
        wb.close()

    def delete(self, num):
        wb = load_workbook('student_scores.xlsx')
        ws = wb.active

        ws.cell(row=int(num) + 1, column=1).value = ""
        ws.cell(row=int(num) + 1, column=2).value = ""
        ws.cell(row=int(num) + 1, column=3).value = ""
        ws.cell(row=int(num) + 1, column=4).value = ""
        ws.cell(row=int(num) + 1, column=5).value = ""
        ws.cell(row=int(num) + 1, column=6).value = ""
        ws.cell(row=int(num) + 1, column=7).value = ""
        ws.cell(row=int(num) + 1, column=8).value = ""
        ws.cell(row=int(num) + 1, column=9).value = ""

        wb.save('student_scores.xlsx')
        wb.close()

    def loadrow(self):
        rows = []

        wb = load_workbook('student_scores.xlsx')
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            rows.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value])
        wb.close()
        return rows


    def createfile(self):
        wb = Workbook()
        ws = wb.active

        ws.append(['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트', '총점', '학점'])

        wb.save("student_scores.xlsx")
        wb.close()
