from openpyxl import *

class MExcel:
    def __init__(self):
        pass

    def save(self,a,b,c):
        wb = load_workbook('mexcel.xlsx')
        ws = wb.active
        ws.append([a,b,c,a+b])
        wb.save('mexcel.xlsx')
        wb.close()

    def load(self):
        wb = load_workbook('mexcel.xlsx')
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end=' ')
            print()

        wb.save('mexcel.xlsx')
        wb.close()

    def create(self):
        wb = Workbook()
        wb.save('mexcel.xlsx')
        wb.close()

