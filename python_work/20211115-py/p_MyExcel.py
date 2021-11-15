
from openpyxl import *


class MExcel:

    def savefn(self, ko, eng, math, tot, avg):
        wb = load_workbook("a.xlsx")
        ws = wb.active
        ws.append([ko, eng, math, tot, avg])
        wb.save('a.xlsx')
        wb.close()

    def loadfn(self):
        wb = load_workbook('a.xlsx')
        ws = wb.active

        ws.delete_rows()

        wb.save('a.xlsx')
        wb.close()
        print("MyExcel savefn")

    def createfn(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['국어', '영어', '수학', '총점', '평균'])
        wb.save('a.xlsx')
        wb.close()
