import time
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PyQt5 import uic, QtGui, QtWidgets, QtCore
import smtplib
from imap_tools import MailBox
from account_PC import * #<- 이걱 각자 이메일 주소,비번 저장된걸로...
from openpyxl import *

from email.message import EmailMessage


class Customer_Email:

    def __init__(self):
        pass

    def Load_List(self):
        rows = []
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active
        for row in ws.iter_rows():
            rows.append(row[0].value)


        wb.close()
        return rows

    def RemoveEmail(self, email):
        rows = []
        count = 1
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            rows.append(row[0].value)

        for e in rows:
            if email != e:
                count += 1
            elif email == e:
                ws.delete_rows(count)

        wb.save("Customer_Mail.xlsx")
        wb.close()

    def SaveEmail(self, email):
        rows = []
        wb = load_workbook("Customer_Mail.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            rows.append(row[0].value)
        if len(rows) >= 1:
            for i in range(0, len(rows)):
                if rows[i] == email:
                    return
        ws.append([email,1])
        wb.save("Customer_Mail.xlsx")
        wb.close()

    def CreateEmail_xl(self):
        wb = Workbook()  # 새 워크북 생성
        ws = wb.active  # 현재 활성화된 sheet 가져옴
        ws.append(['Email List','subscribe'])

        ws.title = "emal"  # sheet 의 이름을 변경
        wb.save("Customer_Mail.xlsx")
        wb.close()

    def recive_mail(self):#류경문 - 이메일 받기
        wb=load_workbook("Customer_Mail.xlsx")
        ws=wb.active

        mailbox = MailBox("imap.gmail.com", 993)
        mailbox.login(EMAIL_ADDRESS, EMAIL_PASSWORD, initial_folder="INBOX")
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

        for msg in mailbox.fetch(('UNSEEN TEXT "AI04B"'),limit=5, reverse=False):

            if msg.from_ is not None:#이메일 내용에 등록 이란 단어가 포함되어있으면 엑셀에 1을 추가
                if "등록" in msg.text:

                    ws.append([msg.from_,1])
                    print("새로 등록한 고객",msg.from_)

                if "해지" in msg.text: #내용에 해지란 단어가 있으면 0을 추가

                    ws.append([msg.from_, 0])
                    print("해지 고객",msg.from_)

                    cu_address=msg.from_ #해지 안내 메일 발송
                    msg = EmailMessage()
                    msg["Subject"] = cu_address + "님 해지완료. "
                    msg["From"] = EMAIL_ADDRESS
                    msg["TO"] = cu_address
                    msg.set_content(cu_address+ "고객님 정상 해지 완료! \n감사합니다.")

                    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
                        smtp.ehlo()
                        smtp.starttls()
                        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

                        smtp.send_message(msg)

                else:
                    self.del_list()#이메일 중복 삭제,해지 요청한사람 삭제 작업으로 넘김

        wb.save("Customer_Mail.xlsx")
        wb.close()

        self.del_list()#이메일 중복 삭제,해지 요청한사람 삭제 작업으로 넘김

    def del_list(self):#류경문 - 중복제거 + 해지 고객 삭제
        del_same = pd.read_excel("Customer_Mail.xlsx", sheet_name="emal")
        df = del_same.drop_duplicates(["Email List"], keep="last")

        df.to_excel("Customer_Mail.xlsx", sheet_name="emal", index=False)

        del_same = pd.read_excel("Customer_Mail.xlsx", sheet_name="emal")
        df1 = del_same[del_same['subscribe'] == 1] #B열에 숫자 1이 없는 행은 전부 삭제
                                                   #해지 요청한 사용자는 0이 들어가서 이단계에서 삭제됨

        df1.to_excel("Customer_Mail.xlsx", sheet_name="emal", index=False)
        print("중복,해지고객 관리 완료")

    def send_mail(self): #류경문 - 이메일 보내기

        wb = load_workbook('Customer_Mail.xlsx')#현재 저장된 고객 명단
        ws = wb.active

        wb1 = load_workbook("검색결과.xlsx",data_only=True)#쿠팡 검색 결과
        ws1 = wb1.active

        for temp1 in ws.iter_rows(min_row=2):#고객 명단 불러오기
            customer_addr =temp1[0].value

            for i in ws1.iter_rows(min_row=2):#검색 결과 불러오기
                img_tag = i[5].value


            s = smtplib.SMTP('smtp.gmail.com', 587)
            s.starttls()
            s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            msg = MIMEMultipart('alternative')
            msg['Subject'] = customer_addr+"고객님만을 위한 특가!"
            msg['From'] = EMAIL_ADDRESS
            msg['To'] = customer_addr

            body = '''\n<h2>더이상 연락을 받기 싫으신분은 답장하기를 눌러서 보내기를 누르시면 해지 처리 해 드립니다.  \n
                   AI04B</h2>\n
                   <a href="'''+i[4].value+''''">쿠팡 바로가기</a>
            ''' + img_tag + '''

                 '''
            #body 안에 HTML 작성가능(img_tag에는 웹스크래핑으로 얻은 html 태그가 담겨져있음)

            body = (body.replace('[', '').replace(',',''))
            text = MIMEText(body, 'html', 'utf-8')
            msg.attach(text)

            wb.close()
            s.sendmail(EMAIL_ADDRESS, customer_addr , msg.as_string())
            s.quit()
            print(customer_addr +"님에게 발송 완료")
            wb1.close()
            wb.close()








