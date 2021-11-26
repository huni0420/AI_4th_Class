import sys
import time
import pyautogui
from PyQt5.QtWidgets import *
import pyperclip

from openpyxl import Workbook

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()


    def doms(self):
        pyautogui.hotkey("win", "r")
        pyautogui.write("mspaint")
        pyautogui.press("enter")

        time.sleep(3)

        box = pyautogui.locateOnScreen('paint.PNG', confidence=0.8)
        pyautogui.click(box)
        pyautogui.click()
        pyautogui.move(10, 200, duration=1)
        pyautogui.click()

        time.sleep(2)

        t = "참잘했어요"
        pyperclip.copy(t)
        pyautogui.hotkey("ctrl", "v")

        pyautogui.sleep(2)
        mywindow = pyautogui.getActiveWindow()
        mywindow.close()

        time.sleep(1)
        pyautogui.write('n')

        # pyautogui.hotkey("alt", "f4")
        # pyautogui.press("n")

        wb = Workbook()
        ws = wb.active

        count = ws['A1'].value
        if count == None:
            count = '1'
        else:
            count = int(count)+1


        ws['A1'] = count

        ws['B1'] = count + 'did'

        wb.save('log.xlsx')
        wb.close()

    def initUI(self):
        self.mpaintbtn = QPushButton("그림판", self)
        self.mpaintbtn.move(30, 20)
        self.mpaintbtn.clicked.connect(self.doms)
        self.setWindowTitle('First Application')
        self.move(300, 300)
        self.resize(400, 200)
        self.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
